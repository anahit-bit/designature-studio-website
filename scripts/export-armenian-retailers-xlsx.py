# -*- coding: utf-8 -*-
"""Download the LIVE Armenian-retailer directory from Sanity into an .xlsx.
Reflects exactly what's in Studio right now (incl. anyone's edits)."""
import json, io, urllib.request
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT = '305mgeeu'
GROQ = ('*[_type=="armenianRetailer"]{nameEN,nameAM,"slug":slug.current,category,tags,'
        'budget,collabClass,status,verifiedAt,deal,notes,website,instagram,facebook,'
        'contact,phone,email,address,description,featured,order} | order(order asc, nameEN asc)')
url = f'https://{PROJECT}.api.sanity.io/v2024-01-01/data/query/production?query=' + urllib.parse.quote(GROQ)
OUT = r'G:/My Drive/Designature Studio/Retail/Armenian-Retailers-from-Sanity.xlsx'

data = json.load(urllib.request.urlopen(url, timeout=30))['result']

COLS = ['nameEN','nameAM','category','tags','budget','collabClass','status','verifiedAt',
        'deal','notes','website','instagram','facebook','contact','phone','email','address','description']
HEAD = {'nameEN':'Name (English)','nameAM':'Name (Armenian)','category':'Category','tags':'Tags',
        'budget':'Budget','collabClass':'Class','status':'Status','verifiedAt':'Verified on',
        'deal':'Deal / cashback','notes':'Notes / ships from','website':'Website','instagram':'Instagram',
        'facebook':'Facebook','contact':'Contact','phone':'Phone','email':'Email','address':'Address',
        'description':'Description'}

wb = openpyxl.Workbook(); sh = wb.active; sh.title = 'Retailers (live from Sanity)'
navy = PatternFill('solid', fgColor='0B2240'); hf = Font(color='FFFFFF', bold=True, size=10)
thin = Side(style='thin', color='DDDDDD'); border = Border(left=thin,right=thin,top=thin,bottom=thin)
for j,c in enumerate(COLS,1):
    cell = sh.cell(1,j,HEAD[c]); cell.fill=navy; cell.font=hf
    cell.alignment=Alignment(horizontal='center',vertical='center')
for i,r in enumerate(sorted(data,key=lambda x:(x.get('order') or 9999, x.get('nameEN') or '')),2):
    for j,c in enumerate(COLS,1):
        v = r.get(c)
        if isinstance(v,list): v = ', '.join(v)
        cell = sh.cell(i,j,v if v is not None else '')
        cell.alignment=Alignment(vertical='top',wrap_text=(c in ('deal','notes','tags','description')))
        cell.border=border
W={'nameEN':22,'nameAM':18,'category':18,'tags':24,'budget':9,'collabClass':8,'status':12,'verifiedAt':12,
   'deal':28,'notes':38,'website':26,'instagram':26,'facebook':22,'contact':22,'phone':20,'email':22,'address':22,'description':32}
for j,c in enumerate(COLS,1): sh.column_dimensions[openpyxl.utils.get_column_letter(j)].width=W[c]
sh.freeze_panes='A2'
n=len(data)+1
def dv(f): d=DataValidation(type='list',formula1=f,allow_blank=True); sh.add_data_validation(d); return d
dv('"Furniture,Lighting,Doors,Windows,Flooring,Tiles & Ceramics,Curtains & Textiles,Stone,Wall Panels & Decor,Electrical & Switches,Sanitaryware,Glass & Mirrors,Paint & Coatings,Building Materials,Heating & HVAC,Kitchen & Bath,Accessories & Decor,Metalwork,Plants & Greenery,Smart Home"').add(f'C2:C{n}')
dv('"low,mid,high"').add(f'E2:E{n}')
dv('"A,B,C,unsorted"').add(f'F2:F{n}')
dv('"active,unverified,closed"').add(f'G2:G{n}')
wb.save(OUT)
print(f'Exported {len(data)} vendors from Sanity -> {OUT}')
