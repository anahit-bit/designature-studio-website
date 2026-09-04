"""
Compile the owner's workbook into the text the image models actually receive.

    python scripts/aivision/compile-rulebook.py

Reads   E:\\Business\\Claude\\_Plan\\Website\\AI-Vision-Rulebook.xlsx
Writes  services/aiVision/rulebook.generated.ts

The owner edits the workbook; the code imports the generated constants. Nobody
hand-writes prompt text in code — that is how the rulebook and the prompt drift
apart.

Sheets consumed:
  Rules          "Prompt text" is the exact sentence sent; blank = never sent
                 (method rules like RD21-RD23 are for us, not the model).
                 "Engines" = gemini | staging | both | none.
                 "Prompt section" = clear | architecture | programme | furnish |
                 finishes —
                 which step of the Gemini prompt the rule is rendered under.
                 One flat block was how RD8 (clear the room) came to sit inside
                 a heading that said CRITICAL ARCHITECTURAL CONSTRAINTS, among
                 thirteen rules that all say "keep": the model read the heading
                 and kept the furniture (2026-09-04).
  Style Briefs   the seven sections, reassembled into one numbered brief.
  Palettes       nine colours per style; the accent-role ones are the pool a
                 single per-generation accent is drawn from.
  Room Programs  what furniture each room type must contain.
  Paint 2026     the Colour-of-the-Year modifier that overrides the accent.

The staging rule set is deliberately tiny. A staging model EDITS the photo it is
given, and a long prompt makes it ignore that photo and draw a fresh room: on
2026-08-29 the 500-word staging prompt scored 0/16 with outputs unrelated to the
source, and the 111-word version scored 8/16. Length there is drift, not
thoroughness (RD23).
"""
import io
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

XLSX = r"E:\Business\Claude\_Plan\Website\AI-Vision-Rulebook.xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.normpath(os.path.join(HERE, "..", "..", "services", "aiVision", "rulebook.generated.ts"))

VALID_ENGINES = {"gemini", "staging", "both", "none"}
VALID_ROLES = {"field", "neutral", "accent"}
# Order matters: this is the order the steps appear in the generated prompt.
PROMPT_SECTIONS = ("clear", "architecture", "programme", "furnish", "finishes")
SECTION_TITLES = [
    "COLOR PALETTE", "MATERIALS & FINISHES", "FURNITURE CHARACTER", "LIGHTING",
    "WALL & CEILING TREATMENT", "DECOR & STYLING", "OVERALL MOOD",
]


def esc(s: str) -> str:
    """Escape for a TS template literal."""
    return s.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")


def cols_of(ws):
    header = [(c.value or "").strip() if isinstance(c.value, str) else "" for c in ws[1]]
    return {name: i for i, name in enumerate(header)}


def need(ws, colmap, sheet, *names):
    for n in names:
        if n not in colmap:
            sys.exit(f"'{sheet}' sheet is missing the '{n}' column.")


def main() -> None:
    if not os.path.exists(XLSX):
        sys.exit(f"Workbook not found: {XLSX}\nSeed it with scripts/aivision/seed-workbook.py")
    wb = load_workbook(XLSX, data_only=True)
    for sheet in ("Rules", "Style Briefs", "Palettes", "Room Programs", "Paint 2026"):
        if sheet not in wb.sheetnames:
            sys.exit(f"Workbook is missing the '{sheet}' sheet.")

    # ── Rules ───────────────────────────────────────────────────────────────
    ws = wb["Rules"]
    col = cols_of(ws)
    need(ws, col, "Rules", "ID", "Level", "Status", "Prompt text", "Engines", "Prompt section")
    gemini, staging, skipped, removed = [], [], [], []
    sections = {name: [] for name in PROMPT_SECTIONS}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["ID"]]:
            continue
        rid = str(row[col["ID"]]).strip()
        status = str(row[col["Status"]] or "").strip().lower()
        text = str(row[col["Prompt text"]] or "").strip()
        engines = str(row[col["Engines"]] or "none").strip().lower()
        level = str(row[col["Level"]] or "").strip()
        if status == "removed":
            removed.append(rid)
            continue
        if engines not in VALID_ENGINES:
            sys.exit(f"{rid}: Engines must be one of {sorted(VALID_ENGINES)} — got '{engines}'.")
        if not text or engines == "none":
            skipped.append(rid)
            continue
        entry = (rid, level, text)
        if engines in ("gemini", "both"):
            gemini.append(entry)
            # A rule the Gemini path sends must say which step it belongs to, or
            # it silently vanishes from the prompt it was written for.
            section = str(row[col["Prompt section"]] or "").strip().lower()
            if section not in PROMPT_SECTIONS:
                sys.exit(
                    f"{rid}: Prompt section must be one of {list(PROMPT_SECTIONS)} "
                    f"for a rule sent to gemini — got '{section}'."
                )
            sections[section].append(entry)
        if engines in ("staging", "both"):
            staging.append(entry)

    empty = [s for s in PROMPT_SECTIONS if not sections[s]]
    if empty:
        sys.exit(f"No rules assigned to prompt section(s): {', '.join(empty)}.")

    # ── Style briefs ────────────────────────────────────────────────────────
    ws = wb["Style Briefs"]
    col = cols_of(ws)
    need(ws, col, "Style Briefs", "Preset key")
    section_cols = [i for name, i in sorted(col.items(), key=lambda kv: kv[1])
                    if re.match(r"^\d\s*·", name)]
    if len(section_cols) != 7:
        sys.exit(f"'Style Briefs' must have 7 numbered section columns — found {len(section_cols)}.")
    briefs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["Preset key"]]:
            continue
        key = str(row[col["Preset key"]]).strip()
        parts = []
        for n, ci in enumerate(section_cols):
            body = str(row[ci] or "").strip()
            if not body:
                sys.exit(f"{key}: section {n + 1} is empty. A brief with a hole in it produces a vague room.")
            parts.append(f"{n + 1}. {SECTION_TITLES[n]}: {body}")
        briefs[key] = "\n".join(parts)

    # ── Palettes ────────────────────────────────────────────────────────────
    ws = wb["Palettes"]
    col = cols_of(ws)
    need(ws, col, "Palettes", "Preset key", "Colour name", "Hex", "Role")
    palettes, current = {}, None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        key = str(row[col["Preset key"]] or "").strip()
        if key:
            current = key
        if not current:
            continue
        name = str(row[col["Colour name"]] or "").strip()
        hexv = str(row[col["Hex"]] or "").strip().upper()
        role = str(row[col["Role"]] or "").strip().lower()
        if not name:
            continue
        if not re.fullmatch(r"#[0-9A-F]{6}", hexv):
            sys.exit(f"{current} / {name}: '{hexv}' is not a #RRGGBB hex.")
        if role not in VALID_ROLES:
            sys.exit(f"{current} / {name}: Role must be one of {sorted(VALID_ROLES)} — got '{role}'.")
        palettes.setdefault(current, []).append((name, hexv, role))

    for key, colours in palettes.items():
        accents = [c for c in colours if c[2] == "accent"]
        if len(accents) < 3:
            sys.exit(f"{key}: needs at least 3 accent colours or its generations stop varying "
                     f"— found {len(accents)}.")

    missing_pal = sorted(set(briefs) - set(palettes))
    if missing_pal:
        sys.exit(f"No palette for: {', '.join(missing_pal)}. Every style needs one.")

    # ── Room programs ───────────────────────────────────────────────────────
    ws = wb["Room Programs"]
    col = cols_of(ws)
    need(ws, col, "Room Programs", "Room key", "Programme rule handed to the model")
    programs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["Room key"]]:
            continue
        key = str(row[col["Room key"]]).strip()
        text = str(row[col["Programme rule handed to the model"]] or "").strip()
        if not text:
            sys.exit(f"Room '{key}' has no programme rule — it would generate a living room.")
        programs[key] = text

    # ── Paint 2026 ──────────────────────────────────────────────────────────
    ws = wb["Paint 2026"]
    col = cols_of(ws)
    need(ws, col, "Paint 2026", "Id", "Colour", "Brand / code", "Hex", "Role", "Instruction sent to the model")
    paints = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["Id"]]:
            continue
        hexv = str(row[col["Hex"]] or "").strip().upper()
        if not re.fullmatch(r"#[0-9A-F]{6}", hexv):
            continue  # trailing note rows
        paints.append({
            "id": str(row[col["Id"]]).strip(),
            "name": str(row[col["Colour"]]).strip(),
            "brand": str(row[col["Brand / code"]]).strip(),
            "hex": hexv,
            "role": str(row[col["Role"]] or "accent").strip().lower(),
            "instruction": str(row[col["Instruction sent to the model"]] or "").strip(),
        })

    # ── Emit ────────────────────────────────────────────────────────────────
    def rules_ts(entries):
        return "\n".join(f'  {{ id: "{r}", level: "{l}", text: `{esc(t)}` }},' for r, l, t in entries)

    def block(entries):
        # The rule id rides along in the prompt so a benchmark failure traces
        # straight back to the row that authorised it.
        return "\n".join(f"- [{rid}] {text}" for rid, _l, text in entries)

    sections_ts = "\n".join(
        f'  {name}: `{esc(block(sections[name]))}`,' for name in PROMPT_SECTIONS
    )

    briefs_ts = "\n".join(f'  {k}: `{esc(v)}`,' for k, v in briefs.items())
    programs_ts = "\n".join(f'  {k}: `{esc(v)}`,' for k, v in programs.items())
    palettes_ts = "\n".join(
        f'  {k}: [\n'
        + "\n".join(f'    {{ name: "{n}", hex: "{h}", role: "{r}" }},' for n, h, r in v)
        + "\n  ],"
        for k, v in palettes.items())
    paints_ts = "\n".join(
        f'  {{ id: "{p["id"]}", name: "{esc(p["name"])}", brand: "{esc(p["brand"])}", '
        f'hex: "{p["hex"]}", role: "{p["role"]}", instruction: `{esc(p["instruction"])}` }},'
        for p in paints)

    total_accents = sum(len([c for c in v if c[2] == "accent"]) for v in palettes.values())

    src = f'''/**
 * GENERATED FILE — DO NOT EDIT BY HAND.
 *
 * Source of truth: _Plan\\Website\\AI-Vision-Rulebook.xlsx
 * Regenerate with:  python scripts/aivision/compile-rulebook.py
 *
 * Edit the workbook, re-run the compiler, commit both. Hand-editing this file
 * makes the rulebook and the prompt disagree, which is how a rule quietly stops
 * being a rule.
 *
 * This build: {len(gemini)} gemini rules · {len(staging)} staging · {len(skipped)} not sent to any model{f" · {len(removed)} removed" if removed else ""}
 *             {len(briefs)} style briefs · {len(palettes)} palettes ({total_accents} accents) · {len(programs)} room programs · {len(paints)} paint modifiers.
 */

export interface CompiledRule {{
  id: string;
  level: string;
  text: string;
}}

/** One colour on a style's palette card. `accent` is the pool a generation draws from. */
export interface PaletteColour {{
  name: string;
  hex: string;
  role: "field" | "neutral" | "accent";
}}

/** A Colour-of-the-Year modifier — overrides the accent, never the style. */
export interface PaintModifier {{
  id: string;
  name: string;
  brand: string;
  hex: string;
  role: string;
  instruction: string;
}}

/** Full constraint set for the regenerative (Gemini) path. */
export const GEMINI_RULES: CompiledRule[] = [
{rules_ts(gemini)}
];

/**
 * Deliberately minimal set for the staging (img2img) path — see RD23. A staging
 * model edits the photo it is handed, and a long prompt makes it ignore that
 * photo entirely. Do not pad this list.
 */
export const STAGING_RULES: CompiledRule[] = [
{rules_ts(staging)}
];

/** Render a rule set as prompt bullets, each tagged with its rulebook id. */
export function renderRules(rules: CompiledRule[]): string {{
  return rules.map((r) => `- [${{r.id}}] ${{r.text}}`).join("\\n");
}}

/** Pre-rendered blocks, so the common path costs nothing at request time. */
export const GEMINI_RULES_BLOCK = `{esc(block(gemini))}`;

export const STAGING_RULES_BLOCK = `{esc(block(staging))}`;

/** The order the prompt's steps are written in. */
export type PromptSection = {" | ".join(f'"{s}"' for s in PROMPT_SECTIONS)};

/**
 * The Gemini rules split into the four steps of the prompt, in order.
 *
 * A single block was how "clear the room" (RD8) ended up filed under a heading
 * reading CRITICAL ARCHITECTURAL CONSTRAINTS, next to thirteen rules that say
 * keep this exactly — and the model kept the furniture. Removal is now its own
 * step, stated before the preservation rules rather than inside them.
 */
export const GEMINI_RULE_SECTIONS: Record<PromptSection, string> = {{
{sections_ts}
}};

/** The seven-section description behind every style chip. */
export const STYLE_BRIEFS: Record<string, string> = {{
{briefs_ts}
}};

/** Nine paint colours per style — the card the owner reviews. */
export const STYLE_PALETTES: Record<string, PaletteColour[]> = {{
{palettes_ts}
}};

/** What furniture each room type must contain. */
export const ROOM_PROGRAM_RULES: Record<string, string> = {{
{programs_ts}
}};

/** 2026 Colours of the Year. A modifier, not a style. */
export const PAINT_MODIFIERS: PaintModifier[] = [
{paints_ts}
];
'''

    io.open(OUT, "w", encoding="utf-8", newline="\n").write(src)
    print(f"wrote {OUT}")
    print(f"  gemini   : {len(gemini)} rules ({sum(len(t.split()) for _, _, t in gemini)} words)")
    for name in PROMPT_SECTIONS:
        print(f"    {name:<13}: {len(sections[name])} rules")
    print(f"  staging  : {len(staging)} rules ({sum(len(t.split()) for _, _, t in staging)} words)")
    print(f"  not sent : {len(skipped)} ({', '.join(skipped)})")
    if removed:
        print(f"  removed  : {len(removed)} ({', '.join(removed)})")
    print(f"  briefs   : {len(briefs)}")
    print(f"  palettes : {len(palettes)} styles, {total_accents} accents total")
    print(f"  programs : {len(programs)}")
    print(f"  paints   : {len(paints)}")


if __name__ == "__main__":
    main()
