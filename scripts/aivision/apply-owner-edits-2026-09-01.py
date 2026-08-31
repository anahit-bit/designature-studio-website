"""
Owner review 2026-09-01 — applied to the workbook, which is the source of truth.

  Art Deco    dining should be brushed brass -> promote Brushed Brass to accent
  Industrial  "too much industrial" -> contemporary loft, keep palette, use oxblood
  Japandi     bathroom needs tiles      -> fixed in the BATHROOM room programme (all styles)
  Hallways    "does not lead anywhere"  -> fixed in the HALLWAY room programme (all styles)
  Mid-Century too many fluted walls     -> slats become rare, never in small rooms
  Modern      lacks personality, too few accessories, make it warmer
  Minimalist  add greens; "does not mean nothing at all, just very few things"
  Maximalist  reads the same as Dopamine -> deep jewel + pattern, not candy brights
  Biophilic   more green, bring nature in (moss stays banned)
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

XLSX = r"E:\Business\Claude\_Plan\Website\AI-Vision-Rulebook.xlsx"
TERRA = "9E5E41"
COL = {"colour": 4, "materials": 5, "furniture": 6, "lighting": 7, "walls": 8, "decor": 9, "mood": 10}

# ── Brief edits: (preset, section) -> new text ──────────────────────────────
BRIEFS = {
 ("industrial", "materials"):
   "one wall of exposed brick or raw concrete at most, the rest smooth plaster; blackened steel used as slim frames "
   "rather than bulk; warm reclaimed oak and walnut with visible grain; soft full-grain leather in tan and oxblood, "
   "aged not distressed; heavy linen and wool boucle; wool rugs; matte ceramics; antique brass alongside the steel. "
   "No riveted plate, no cast iron, no visible wiring runs.",
 ("industrial", "furniture"):
   "contemporary silhouettes with an industrial memory rather than factory salvage — a deep comfortable sofa in tan or "
   "oxblood leather, slim steel-framed tables with warm timber tops, a soft upholstered armchair, open shelving in thin "
   "black steel. Comfortable first. No factory carts, no mechanic's stools, no riveted furniture, nothing that looks "
   "like reclaimed machinery.",
 ("industrial", "walls"):
   "ONE feature wall in exposed brick or raw concrete, with the remaining walls in smooth warm plaster or matte paint. "
   "Keep the ceiling clean and flat: no exposed ductwork, no pipe runs, no visible conduit, no cable trays. Industrial "
   "character comes from that single honest surface, not from stripping the whole room back.",
 ("industrial", "decor"):
   "framed abstract or architectural prints in thin black frames, a few well-chosen ceramics, stacked books, a large "
   "leafy plant in a concrete or terracotta planter, a wool throw, a woven rug that softens the floor. No vintage "
   "signage, no gears or tools as ornament, no oversized exposed-mechanism clocks.",
 ("industrial", "mood"):
   "warm, relaxed, contemporary. A loft someone actually lives in — the brick and steel are background, and leather, "
   "timber, wool and plants carry the room. Grounded and easy rather than raw or hard-edged.",

 ("mid_century", "walls"):
   "matte painted walls in warm neutrals across the whole room. Slatted or fluted timber is RARE — at most ONE wall, "
   "only in a generously sized living or dining room, and only when the room has enough width to carry it. Never slat "
   "a hallway, a bathroom, a kitchen or a small room, and never slat more than one surface. Most rooms in this style "
   "have no slatted wall at all. Simple flat ceilings, no retro-era trim.",

 ("modern", "colour"):
   "warm and contemporary rather than cold — Warm White #F4F0E9, Soft Oat #E3D9C9, Warm Taupe #A89F92, Walnut #6B5544, "
   "Charcoal #333333, Muted Olive #7C7F5E, Terracotta #B87A5E. Warm neutrals carry the room and one deeper tone anchors "
   "it. Avoid a pure-white, cool-grey scheme; there should be visible warmth in every image.",
 ("modern", "materials"):
   "warm oak and walnut with visible grain, honed travertine or limestone, matte lacquer (never high-gloss), brushed "
   "brass and blackened steel in small doses, wool and heavy linen upholstery, boucle, leather in tan, matte ceramics, "
   "clear and reeded glass. Texture matters as much as line.",
 ("modern", "decor"):
   "generously styled, not bare — a large artwork plus a smaller framed pair, stacked design books, three or four "
   "ceramic vessels of different heights, a woven bowl, a throw over the sofa arm, cushions in mixed textures, two or "
   "three real plants in matte planters, a tray on the coffee table. Surfaces are curated and lived-in. Clean lines do "
   "NOT mean empty surfaces — this is not minimalism.",
 ("modern", "mood"):
   "calm, warm, contemporary and clearly lived in. Clean architecture softened by warm timber, texture and real "
   "objects. Composed rather than stark, and never gallery-cold or showroom-empty.",

 ("minimalist", "colour"):
   "Warm White #F5F3EE, Off-White #FFFFFF, Soft Warm Grey #E0DBD5, Sage Mist #C3CBBD, Light Concrete #C8C4BE, "
   "Olive Leaf #6E7A5A, Warm Charcoal #4A4542, Raw Oak #C4AE8C. Predominantly warm neutral, with a quiet green "
   "carrying the one note of life in the room.",
 ("minimalist", "decor"):
   "few things, chosen carefully — not an empty room. One large artwork, one ceramic vessel, a small stack of books, a "
   "single real plant or branch arrangement in a simple pot, one folded textile. Roughly three or four deliberate "
   "objects in view, each with clear space around it. Surfaces stay largely open, but the room must read as inhabited "
   "and calm rather than unfinished or abandoned.",
 ("minimalist", "mood"):
   "serene, warm and quietly alive. Restraint, not absence — the eye rests, but something green is growing and the "
   "room feels like somebody's, not a showroom awaiting delivery.",

 ("maximalist", "colour"):
   "deep, rich and saturated rather than bright — Ink Green #1F3D33, Emerald #2A6B4A, Oxblood #6E2A2A, "
   "Peacock Teal #147A80, Ochre Gold #C08A2E, Aubergine #4A2A45, Walnut #5A3E2B, Warm Chalk #F4EFE6. Jewel and "
   "heritage tones layered together. Deliberately NOT a bright candy palette — no bubblegum pink, no sky blue, no "
   "sunflower yellow; those belong to Dopamine, which is a different style.",
 ("maximalist", "materials"):
   "layered and tactile — velvet and mohair in jewel tones, patterned wool rugs stacked over one another, printed "
   "linen and cotton in large-scale florals, stripes and geometrics, lacquered and burled wood, antique brass, marble, "
   "glazed and patterned tile, fringed and tasselled trims. Pattern meeting pattern is the point.",
 ("maximalist", "furniture"):
   "collected rather than bought as a set — a deep velvet sofa in a jewel tone, two armchairs that do not match each "
   "other, an antique or vintage cabinet with real age beside a contemporary table, a patterned ottoman, a bar cart. "
   "Silhouettes are generous and comfortable. The room should look assembled over years by someone with strong taste, "
   "not styled in an afternoon.",
 ("maximalist", "walls"):
   "deep saturated paint on every wall, or a large-scale patterned wallpaper — botanical, damask-scale florals, "
   "chinoiserie-style murals or bold geometrics. A dense floor-to-ceiling gallery wall in mixed antique brass and dark "
   "timber frames. Colour may continue onto the ceiling. Ceilings stay flat.",
 ("maximalist", "decor"):
   "abundance, carefully arranged — a dense gallery wall, books stacked everywhere, collected ceramics and glassware "
   "grouped by colour, table lamps with patterned or pleated shades, layered rugs, cushions in four or five different "
   "prints, tall plants, objects with provenance. Every surface holds something and nothing looks accidental.",
 ("maximalist", "mood"):
   "rich, layered, personal and confident. Depth and pattern rather than brightness — a room that rewards a second "
   "look because there is always something else in it. Warm and enveloping, never loud or childlike.",

 ("biophilic", "colour"):
   "Natural White #F0EBE2, Stone Cream #E4DED5, Sage Green #8FA383, Leaf Green #6B8C5A, Deep Fern #3F5A3A, "
   "Bark Brown #7A5C3A, Warm Teak #A8794C, Terracotta #B86E4A. Green runs right through the room — in the planting "
   "first, and in soft painted or upholstered greens second. Warm neutrals hold the background so the greens read as "
   "living rather than decorative.",
 ("biophilic", "walls"):
   "warm vertical timber battens or slats across one wall, raw lime plaster in earthy tones, exposed natural stone, "
   "timber-slat ceilings. One wall MAY be painted a soft sage or deep fern green where it sits behind planting. MOSS "
   "IS NOT PART OF THIS STYLE — never clad a wall, panel or ceiling in moss. A living vertical garden of real potted "
   "plants is permitted on at most ONE wall of a large, well-daylit room.",
 ("biophilic", "decor"):
   "bring the outside in across the WHOLE room, evenly. Not one hero plant standing alone, and not wall-to-wall "
   "foliage. Distribute greenery at THREE scales: one or two floor-standing plants or a small indoor tree (Japanese "
   "maple, olive, fiddle leaf, rubber plant) in a stone, concrete or terracotta planter; three or four mid-height "
   "plants on furniture, shelves and window sills; and one or two trailing or climbing plants softening a high edge or "
   "the top of a shelf. Roughly eight to twelve plants in a normal room, varied in leaf shape and in shade of green, "
   "each with enough space to be read individually. Around them: a bed of river pebbles where the floor allows, stone "
   "and unglazed bowls, a branch or dried arrangement, botanical art, seed pods and natural specimens. The room should "
   "feel as though nature runs through it — balanced and lived-with, never a single plant placed for the photograph "
   "and never a garden centre.",
 ("biophilic", "lighting"):
   "daylight is the primary material: skylights, clerestories and full-height glazing wherever the room already has "
   "them, and where a window looks out, let real greenery be visible through it so the planting continues outside. "
   "Supplement with warm LED at golden-hour temperature and fixtures in rattan, paper, stone or wood. No cold-white "
   "sources.",
 ("biophilic", "mood"):
   "restorative, green and alive. Nature is not a motif here, it is the subject — daylight, timber, stone and real "
   "growing plants at every level of the room, with the greenery outside answering the greenery inside.",
}

# ── Palette edits: preset -> full replacement (9 rows, order preserved) ─────
PALETTES = {
 "art_deco": [  # Brushed Brass promoted so the dining room can be drawn in it
   ("Cream Ivory", "#F2E9D8", "field"), ("Warm Off-White", "#F5EFE0", "field"),
   ("Warm Charcoal", "#3A3A3A", "neutral"), ("Honed Marble Grey", "#CFC9BF", "neutral"),
   ("Brushed Brass", "#B8935E", "accent"), ("Deep Forest", "#2E4E3D", "accent"),
   ("Warm Terracotta", "#B87A5E", "accent"), ("Muted Burgundy", "#7A3A42", "accent"),
   ("Ink Navy", "#2B3A4A", "accent"),
 ],
 "minimalist": [  # two greens: one quiet neutral, one accent that can carry a room
   ("Off-White", "#FFFFFF", "field"), ("Warm White", "#F5F3EE", "field"),
   ("Soft Warm Grey", "#E0DBD5", "neutral"), ("Sage Mist", "#C3CBBD", "neutral"),
   ("Light Concrete", "#C8C4BE", "neutral"), ("Olive Leaf", "#6E7A5A", "accent"),
   ("Warm Charcoal", "#4A4542", "accent"), ("Raw Oak", "#C4AE8C", "accent"),
   ("Black", "#1A1A1A", "accent"),
 ],
 "maximalist": [  # deep jewel + heritage, so it stops colliding with Dopamine
   ("Warm Chalk", "#F4EFE6", "field"), ("Ink Green", "#1F3D33", "field"),
   ("Antique Brass", "#A8823C", "neutral"), ("Walnut", "#5A3E2B", "neutral"),
   ("Emerald", "#2A6B4A", "accent"), ("Oxblood", "#6E2A2A", "accent"),
   ("Peacock Teal", "#147A80", "accent"), ("Ochre Gold", "#C08A2E", "accent"),
   ("Aubergine", "#4A2A45", "accent"),
 ],
 "biophilic": [  # greener again, but the green arrives as planting, never as moss
   ("Natural White", "#F0EBE2", "field"), ("Stone Cream", "#E4DED5", "field"),
   ("Bark Brown", "#7A5C3A", "neutral"), ("Sage Green", "#8FA383", "neutral"),
   ("Riverstone Grey", "#8C8880", "neutral"), ("Warm Teak", "#A8794C", "accent"),
   ("Terracotta", "#B86E4A", "accent"), ("Leaf Green", "#6B8C5A", "accent"),
   ("Deep Fern", "#3F5A3A", "accent"),
 ],
}

# ── Room programmes (apply to EVERY style) ─────────────────────────────────
PROGRAMS = {
 "bathroom":
   "The room MUST be a fully realized BATHROOM. It MUST be properly TILED: real wall tile behind the vanity and "
   "through the shower or bath area, and tiled or stone flooring, with visible grout lines and a tile format, colour "
   "and laying pattern chosen to suit the target style. Include a vanity with sink(s) and mirror, a toilet, a shower "
   "or bathtub (or both), a glazed shower screen where there is a shower, towel bars or rings with real towels, "
   "sconces or vanity lighting, a bath mat, and appropriate styling. Do NOT leave the walls as bare plaster or paint "
   "alone, and do NOT include living-room furniture, bedroom furniture, or dining tables. Every fixture must be a real "
   "bathroom fixture.",
 "hallway":
   "The room MUST be a fully realized HALLWAY, and it MUST READ AS LEADING SOMEWHERE — never a sealed dead-end box. "
   "The view should carry through to a doorway, a wider opening, a turn, a staircase or a window at the far end, so "
   "the eye travels past the foreground and the space clearly connects two parts of a home. Where a source photograph "
   "already shows such an opening, keep it clear and unobstructed; never invent an opening the source does not show. "
   "Furnish it as a hallway: a narrow console or hall table against one wall, wall art or a gallery arrangement, a "
   "runner rug following the length of the space, wall sconces or pendants, and appropriate styling. Do NOT include "
   "living-room seating groups, beds, or dining tables.",
}

wb = load_workbook(XLSX)

# ── Apply brief edits ───────────────────────────────────────────────────────
sb = wb["Style Briefs"]
rows = {sb.cell(r, 2).value: r for r in range(2, sb.max_row + 1) if sb.cell(r, 2).value}
for (preset, section), text in BRIEFS.items():
    if preset not in rows:
        raise SystemExit(f"unknown preset {preset}")
    sb.cell(rows[preset], COL[section]).value = text
print(f"briefs: {len(BRIEFS)} sections rewritten across "
      f"{len({p for p, _ in BRIEFS})} styles")

# ── Apply palette edits ─────────────────────────────────────────────────────
pal = wb["Palettes"]
starts = {}
cur = None
for r in range(2, pal.max_row + 1):
    k = pal.cell(r, 2).value
    if k:
        cur = str(k).strip()
        starts[cur] = r
for preset, colours in PALETTES.items():
    if preset not in starts:
        raise SystemExit(f"no palette block for {preset}")
    s = starts[preset]
    if len(colours) != 9:
        raise SystemExit(f"{preset}: expected 9 colours, got {len(colours)}")
    for i, (name, hexv, role) in enumerate(colours):
        r = s + i
        pal.cell(r, 3).value = i + 1
        pal.cell(r, 4).value = name
        pal.cell(r, 5).value = hexv
        pal.cell(r, 5).font = Font(size=9, name="Consolas")
        pal.cell(r, 6).value = role
        pal.cell(r, 6).font = Font(size=9, bold=True, color=TERRA) if role == "accent" else Font(size=9)
        pal.cell(r, 7).fill = PatternFill("solid", fgColor=hexv.lstrip("#").upper())
    n_acc = sum(1 for c in colours if c[2] == "accent")
    print(f"palette {preset:<12} -> {n_acc} accents")

# ── Apply room-programme edits ──────────────────────────────────────────────
rp = wb["Room Programs"]
prow = {rp.cell(r, 2).value: r for r in range(2, rp.max_row + 1) if rp.cell(r, 2).value}
for key, text in PROGRAMS.items():
    if key not in prow:
        raise SystemExit(f"no room programme row for {key}")
    rp.cell(prow[key], 5).value = text
    print(f"programme {key} rewritten")

wb.save(XLSX)
print("saved", XLSX)
