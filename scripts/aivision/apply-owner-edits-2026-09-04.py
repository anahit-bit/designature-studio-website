# -*- coding: utf-8 -*-
"""
Rulebook edit, 2026-09-04 - the room programme stops building architecture.

Owner reported a Hallway + Trend 2026 concept that cut an archway with a
staircase into a sealed dead-end wall and kept the original bed frame. The
prompt had ordered both: the hallway programme said the space "MUST READ AS
LEADING SOMEWHERE ... a staircase", which the model obeyed over RD3/RD7, and
RD8 (clear the room) was buried inside a block headed CRITICAL ARCHITECTURAL
CONSTRAINTS among thirteen rules that all say "keep".

Three changes:
  1. A "Prompt section" column, so the compiler can order the prompt as steps
     (clear -> architecture -> furnish -> finishes) instead of one flat wall of
     rules with the removal rule lost inside it.
  2. RD24 - the photograph outranks the room programme. Rendered immediately
     above the programme, where the contradiction was being resolved wrongly.
     RD25/RD26 record the two enforcement escalations RD22 demands.
  3. Every room programme reviewed for the same defect: a programme that
     mandates a feature the photograph may not contain.

Run once:  python scripts/aivision/apply-owner-edits-2026-09-04.py
Then:      python scripts/aivision/compile-rulebook.py
"""
import sys
from openpyxl import load_workbook

XLSX = r"E:\Business\Claude\_Plan\Website\AI-Vision-Rulebook.xlsx"

# -- Which prompt step each rule belongs to ----------------------------------
SECTIONS = {
    "RD1": "architecture", "RD2": "architecture", "RD3": "architecture",
    "RD4": "architecture", "RD5": "architecture", "RD6": "architecture",
    "RD7": "architecture",
    "RD8": "clear", "RD9": "clear", "RD11": "clear",
    "RD10": "furnish", "RD12": "furnish", "RD13": "furnish",
    "RD14": "furnish", "RD15": "furnish", "RD17": "furnish",
    "RD16": "finishes", "RD18": "finishes",
}

NEW_RULES = [
    {
        "ID": "RD24",
        "Group": "Governing",
        "Rule": (
            "The photograph outranks the room programme. A programme says what the finished room "
            "CONTAINS; it may never authorise architecture the photo does not show. Where the two "
            "collide, drop the part of the programme that needs building work."
        ),
        "Level": "ABSOLUTE",
        "Why / evidence": (
            "2026-09-04: the Hallway programme demanded a through-view ('a doorway, a wider opening, "
            "a turn, a staircase'); on a sealed dead-end alcove the model cut an arch with a staircase "
            "into the back wall. The programme's positive imperative beat RD3/RD7's negatives, and the "
            "programme's own 'never invent an opening' clause was a subordinate sentence inside the "
            "same MUST."
        ),
        "Enforced by": (
            "Rendered directly above the room programme in the FURNISH step of the Gemini prompt, "
            "plus RD26's photo-conditional note."
        ),
        "Status": "active",
        "Prompt text": (
            "The room programme that follows lists what the finished room CONTAINS - furniture, "
            "fittings, lighting and styling. It can never authorise building work. If any part of it "
            "would need an opening, doorway, arch, staircase, window, wall, recess or change of level "
            "that this photograph does not already show, DROP that part and furnish what is actually "
            "there. A room that reads as a dead end, or as small, or as awkwardly shaped, is the "
            "correct answer when that is the room in the photograph."
        ),
        "Engines": "gemini",
        "Prompt section": "furnish",
        "Owner note": "",
    },
    {
        "ID": "RD25",
        "Group": "Enforcement",
        "Rule": (
            "Openings are counted after generation, not merely forbidden before it. If the output "
            "shows more windows or doors than the source, the image is regenerated once with a "
            "targeted correction."
        ),
        "Level": "HARD",
        "Why / evidence": (
            "RD22 escalation for RD1/RD2/RD7. The AI-029 verify-and-retry only ever measured window "
            "WIDTH, and derived that from the largest window - so a photograph with no window "
            "(hallway, alcove, interior bathroom) ran with no post-generation check at all. That is "
            "exactly the photo that failed on 2026-09-04."
        ),
        "Enforced by": (
            "services/aiVision/imageGeneration.ts - countOpenings() diff against the source "
            "structure, one corrective retry."
        ),
        "Status": "active",
        "Prompt text": "",
        "Engines": "none",
        "Prompt section": "",
        "Owner note": "",
    },
    {
        "ID": "RD26",
        "Group": "Enforcement",
        "Rule": (
            "A room programme is issued against the measured photograph. Where the analysis "
            "contradicts a programme's assumption, a photo-specific note is appended naming what "
            "this room does not have."
        ),
        "Level": "HARD",
        "Why / evidence": (
            "Generic programmes cannot know that this hallway has no doorway or that this bedroom's "
            "only wall carries the window. Stating the absence explicitly is stronger than a general "
            "prohibition."
        ),
        "Enforced by": "services/aiVision/promptTemplates.ts - renderProgrammeNote(roomType, structure).",
        "Status": "active",
        "Prompt text": "",
        "Engines": "none",
        "Prompt section": "",
        "Owner note": "",
    },
]

# -- Room programmes, all ten reviewed ---------------------------------------
PROGRAMS = {
    # REWRITTEN. The through-view was mandatory; it is now the photograph's to give.
    "hallway": (
        "The room MUST be a fully realized HALLWAY - a circulation space, furnished lightly, with the "
        "centre of the floor left clear enough to walk through. Furnish it as a hallway: a narrow "
        "console or hall table against one wall, wall art or a gallery arrangement, a runner rug "
        "following the length of the space, wall sconces or pendants, and a slim bench, hooks or shoe "
        "storage where a wall allows. THE THROUGH-VIEW IS THE PHOTOGRAPH'S TO GIVE, NOT YOURS TO "
        "CREATE. Where the photo already shows a doorway, opening, turn or staircase, keep it exactly "
        "as it is, clear and unobstructed, and let the eye travel to it. Where the photo does not show "
        "one, this hallway ENDS at the wall in front of you: do NOT cut an opening, arch, doorway or "
        "staircase into any wall, and do not suggest one with a recess, a painted arch or a lit "
        "alcove. A hallway that reads as a dead end is the correct result for a dead-end photograph. "
        "Do NOT include living-room seating groups, beds, or dining tables."
    ),
    # TIGHTENED. The fixture list was unconditional, so a small WC had to invent a
    # tiled recess to hold a bath. Tiling stays - flat cladding is a finish (RD16).
    "bathroom": (
        "The room MUST be a fully realized BATHROOM. It MUST be properly TILED: real wall tile behind "
        "the vanity and through the shower or bath area, and tiled or stone flooring, with visible "
        "grout lines and a tile format, colour and laying pattern chosen to suit the target style. "
        "Tile is a finish laid flat onto the walls that are already there - it never changes their "
        "plane. Include a vanity with sink(s) and mirror, a toilet, towel bars or rings with real "
        "towels, sconces or vanity lighting, a bath mat, and appropriate styling. For bathing, KEEP "
        "WHAT THE PHOTOGRAPH HAS: a bath stays a bath, a shower stays a shower, in the same place at "
        "the same size. If the photograph shows neither, fit only what the visible floor area holds "
        "without moving a wall - and if it holds neither, a vanity, toilet and mirror are a complete "
        "answer. Add a glazed shower screen only where there is a shower. Do NOT carve a new recess, "
        "niche, wet-room enclosure or partition to house a fixture, and do NOT enlarge the room to fit "
        "one. Do NOT leave the walls as bare plaster or paint alone, and do NOT include living-room "
        "furniture, bedroom furniture, or dining tables. Every fixture must be a real bathroom fixture."
    ),
    # TIGHTENED. An extractor hood is the classic route to an invented bulkhead
    # (the 2026-07-13 kitchen soffit), and the island was the only conditional item.
    "kitchen": (
        "The room MUST be a fully realized KITCHEN. Include base and wall cabinetry, a countertop with "
        "backsplash, a range or cooktop, a sink with faucet, and the appliances the space takes "
        "(fridge, oven), plus open shelving or a hutch. Fit the run to the walls the photograph "
        "actually shows: cabinetry stands flush against them and stops where they stop. A range hood "
        "is an APPLIANCE - mount it on the existing wall or hang it from the existing flat ceiling, "
        "and never box it into a new bulkhead, soffit, chimney breast or dropped ceiling section. Add "
        "a kitchen island with counter stools ONLY where the visible floor area clearly leaves a "
        "walking route around it; in a narrow or single-wall kitchen, leave it out. Do NOT include "
        "living-room furniture (sofa, armchair, coffee table), bedroom furniture, or dining tables "
        "(unless a small breakfast nook is clearly the intent)."
    ),
    # TIGHTENED. 'nightstands on both sides' is impossible once RD15 pushes the bed
    # against a side wall, and the model resolved that by moving the window instead.
    "bedroom": (
        "The room MUST be a fully realized BEDROOM. The visual anchor is a fully-made bed with a "
        "bedside table and lamp on each side WHERE THE WALL LENGTH ALLOWS - one side only, or a "
        "wall-mounted reading light, is the right answer in a narrow room, and is always better than "
        "moving the bed off its wall or shrinking anything around it. Add a bench at the foot, a "
        "dresser or wardrobe, a rug under the bed, art on the wall, and appropriate styling, each "
        "sized to the room the photograph shows. Do NOT include living-room furniture (sofa as primary "
        "piece, TV area), dining tables, or kitchen cabinetry. A reading chair in the corner is fine "
        "if the room is large. BED PLACEMENT (critical): every existing window or balcony door must "
        "stay fully visible, unobstructed, and exactly where it is - never cover it with the "
        "headboard, and never remove, shrink, move, or replace a window to make room for the bed. If "
        "the only large wall in view holds a window or balcony door, place the bed against a side wall "
        "or offset to one side of the window so the glazing stays completely clear; do not build a new "
        "solid headboard wall over a window, and do not hang art over a window."
    ),
    # TIGHTENED. Bunk beds in a low or small room were an invitation to raise the ceiling.
    "kids_room": (
        "The room MUST be a fully realized KIDS' BEDROOM or PLAYROOM. Include a child-scale bed, a "
        "small desk, storage bins or cubbies, a rug for floor play, playful wall art, and "
        "age-appropriate styling. Every piece is sized to the room in the photograph: use bunk beds "
        "only where the existing ceiling height clearly takes them, and never raise the ceiling, widen "
        "the room or borrow floor area to fit a piece in. Do NOT include adult living-room furniture, "
        "formal dining, or kitchen fixtures. BED PLACEMENT (critical): every existing window or "
        "balcony door must stay fully visible, unobstructed, and exactly where it is - never cover it "
        "with the headboard, and never remove, shrink, move, or replace a window to make room for the "
        "bed. If the only large wall in view holds a window or balcony door, place the bed against a "
        "side wall or offset to one side of the window so the glazing stays completely clear."
    ),
    # TIGHTENED. The pendant is legal (RD18) but was silent about what it hangs from.
    "dining_room": (
        "The room MUST be a fully realized DINING ROOM. The visual anchor is a dining table with 4-8 "
        "matching dining chairs, positioned as the centerpiece under a hanging pendant or chandelier - "
        "hung from the existing flat ceiling, never from a new coffer, tray, dropped section or "
        "bulkhead, and sized so the chairs pull out within the floor the photograph shows. Add a "
        "sideboard or credenza against one wall, wall art, and appropriate styling. Do NOT include "
        "lounge or living-room seating groups, lounge armchairs arranged around a low central table, "
        "media consoles, or televisions, even if the target style is often shown in a living room. Any "
        "style furniture examples below should be REINTERPRETED as dining-room equivalents - a dining "
        "chair in that style, a dining table in that style, a sideboard in that style."
    ),
    # TIGHTENED. Reviewed and left almost as-is; only the sizing clause is new.
    "living_room": (
        "The room MUST be a fully realized LIVING ROOM. Include ONLY furniture appropriate to a living "
        "room - a sofa, one or two armchairs, a coffee table, side tables, floor or table lamps, a "
        "rug, wall art, plants, and appropriate styling, every piece sized to the floor area the "
        "photograph actually shows. In a small room, fewer and smaller pieces are the right answer; "
        "never widen the room to fit the furniture. Do NOT include dining tables, beds, kitchen "
        "cabinetry, desks, or bathroom fixtures."
    ),
    # TIGHTENED. Same sizing clause; the storage rule defers to RD14.
    "home_office": (
        "The room MUST be a fully realized HOME OFFICE. The visual anchor is a desk with a task chair. "
        "Add a task lamp, bookshelves or storage standing flush against an existing wall, monitor(s) "
        "or a laptop, wall art, and appropriate styling, each sized to the room the photograph shows. "
        "Do NOT include living-room furniture as the primary piece, beds, or dining tables. A small "
        "accent chair for a reading corner is fine where the floor area allows."
    ),
    # TIGHTENED. The one programme whose subject IS a structure, so it needed the
    # structure named explicitly as fixed.
    "outdoor": (
        "The room MUST be a fully realized OUTDOOR SPACE (patio, terrace, or balcony as appropriate to "
        "the original photo). Include outdoor-rated seating (sofa, chairs, or dining set as fits the "
        "space), an outdoor rug, planters with real outdoor plants, string lights or outdoor sconces, "
        "and appropriate styling. All materials must be weather-appropriate. Keep the existing "
        "structure - railings, parapets, canopies, posts and the building wall behind - exactly as "
        "photographed; add no pergola, roof, screen wall or built structure the photo does not show. "
        "Do NOT include indoor furniture that would not survive weather."
    ),
    # UNCHANGED - already the best-written programme in the book, and the model the
    # others were rewritten against: it names its own architectural prohibitions.
    "living_dining": None,
}


def main() -> None:
    wb = load_workbook(XLSX)

    # -- Rules: add the Prompt section column --------------------------------
    ws = wb["Rules"]
    header = [(c.value or "").strip() if isinstance(c.value, str) else "" for c in ws[1]]
    if "Prompt section" not in header:
        # Insert just after Engines so it reads next to the routing columns.
        at = header.index("Engines") + 2
        ws.insert_cols(at)
        cell = ws.cell(row=1, column=at, value="Prompt section")
        left = ws.cell(row=1, column=at - 1)
        cell.font = left.font.copy()
        cell.fill = left.fill.copy()
        cell.alignment = left.alignment.copy()
        ws.column_dimensions[cell.column_letter].width = 16
        header = [(c.value or "").strip() if isinstance(c.value, str) else "" for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(header)}

    filled = 0
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=col["ID"]).value
        if not rid:
            continue
        rid = str(rid).strip()
        if rid in SECTIONS:
            ws.cell(row=r, column=col["Prompt section"], value=SECTIONS[rid])
            filled += 1
    print("Rules: prompt section set on %d rows" % filled)

    existing = {
        str(ws.cell(row=r, column=col["ID"]).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=col["ID"]).value
    }
    added = 0
    for rule in NEW_RULES:
        if rule["ID"] in existing:
            print("Rules: %s already present, left alone" % rule["ID"])
            continue
        r = ws.max_row + 1
        for name, value in rule.items():
            if name in col:
                ws.cell(row=r, column=col[name], value=value)
        added += 1
        print("Rules: added %s" % rule["ID"])
    print("Rules: %d new rows" % added)

    # -- Room programmes -----------------------------------------------------
    ws = wb["Room Programs"]
    header = [(c.value or "").strip() if isinstance(c.value, str) else "" for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(header)}
    seen = set()
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=col["Room key"]).value
        if not key:
            continue
        key = str(key).strip()
        seen.add(key)
        if key not in PROGRAMS:
            print("Room Programs: %s not in this edit - left alone" % key)
            continue
        text = PROGRAMS[key]
        if text is None:
            print("Room Programs: %s reviewed, unchanged" % key)
            continue
        ws.cell(row=r, column=col["Programme rule handed to the model"], value=text)
        print("Room Programs: %s rewritten (%d chars)" % (key, len(text)))
    missing = sorted(k for k in PROGRAMS if k not in seen)
    if missing:
        sys.exit("Room keys in this edit that the sheet does not have: %s" % missing)

    wb.save(XLSX)
    print("\nSaved %s" % XLSX)


if __name__ == "__main__":
    main()
