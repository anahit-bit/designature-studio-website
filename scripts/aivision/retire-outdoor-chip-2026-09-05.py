# -*- coding: utf-8 -*-
"""
Retire the Outdoor room chip, 2026-09-05.

RD19 has said outdoor and open-air spaces are OUT OF SCOPE since 2026-08-29
(C062 and C076: both open-air kiosks were replaced by entirely new structures).
The chip stayed live in the room picker anyway, so the rulebook and the product
disagreed about what the tool does. The owner resolved it in favour of the
rulebook.

What this does NOT do: remove the `outdoor` RoomType or its programme. An older
saved concept, or an API caller passing "Outdoor", must still resolve to the
outdoor programme rather than silently becoming a living room - which is exactly
the class of bug the 2026-09-04 work was about.

"Live in UI?" on the Room Programs sheet is the source of truth. The compiler
emits it as LIVE_ROOM_CHIPS and stylePresets.test asserts the component's chip
array matches, so the two cannot drift apart again.

Requires the workbook to be CLOSED in Excel.

Run once:  python scripts/aivision/retire-outdoor-chip-2026-09-05.py
Then:      python scripts/aivision/compile-rulebook.py
"""
import sys
from openpyxl import load_workbook

XLSX = r"E:\Business\Claude\_Plan\Website\AI-Vision-Rulebook.xlsx"

RD19_ENFORCED_BY = (
    "The Outdoor chip was removed from the room picker on 2026-09-05, so the programme can no "
    "longer be requested from the UI. The room type and its programme are retained so an older "
    'saved concept, or an API caller passing "Outdoor", still resolves to the right programme '
    "instead of silently becoming a living room. \"Live in UI?\" on the Room Programs sheet is the "
    "source of truth; the compiler emits it as LIVE_ROOM_CHIPS and stylePresets.test asserts the "
    "chip array matches it."
)


def main() -> None:
    try:
        wb = load_workbook(XLSX)
    except PermissionError:
        sys.exit("Workbook is open in Excel. Close it and run this again.")

    ws = wb["Room Programs"]
    header = [(c.value or "").strip() if isinstance(c.value, str) else "" for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(header)}
    hit = False
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=col["Room key"]).value or "").strip() == "outdoor":
            ws.cell(row=r, column=col["Live in UI?"], value="no")
            hit = True
            print("Room Programs: outdoor -> Live in UI? = no")
    if not hit:
        sys.exit("No 'outdoor' row on the Room Programs sheet.")

    ws = wb["Rules"]
    header = [(c.value or "").strip() if isinstance(c.value, str) else "" for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(header)}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=col["ID"]).value or "").strip() == "RD19":
            ws.cell(row=r, column=col["Enforced by"], value=RD19_ENFORCED_BY)
            print("Rules: RD19 enforcement recorded")

    try:
        wb.save(XLSX)
    except PermissionError:
        sys.exit("Workbook is open in Excel. Close it and run this again.")
    print("\nSaved %s" % XLSX)


if __name__ == "__main__":
    main()
