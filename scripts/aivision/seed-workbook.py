"""
Seed the ONE AI Vision workbook — _Plan\\Website\\AI-Vision-Rulebook.xlsx.

    python scripts/aivision/seed-workbook.py            # refuses if it exists
    python scripts/aivision/seed-workbook.py --force    # OVERWRITES OWNER EDITS

Merges what used to be two files:
  Redesign-My-Room-Rulebook.xlsx  (Rules)          -> "Rules" sheet
  AI-Vision-Style-Briefs.xlsx     (derived view)   -> "Style Briefs" / "Room Programs"
and adds two new sheets, "Palettes" and "Paint 2026".

SEEDING ONLY. It reads the current code to build the first version; after that the
WORKBOOK is the source of truth and `compile-rulebook.py` pushes it back into code.
Running this again with --force throws away whatever the owner has since written,
which is why the guard exists.
"""
import os
import re
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palettes import PALETTES, PAINT_2026  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.normpath(os.path.join(HERE, "..", ".."))
PLAN = r"E:\Business\Claude\_Plan\Website"
OUT = os.path.join(PLAN, "AI-Vision-Rulebook.xlsx")
OLD_RULES = os.path.join(PLAN, "Redesign-My-Room-Rulebook.xlsx")

force = "--force" in sys.argv
if os.path.exists(OUT) and not force:
    sys.exit(f"{OUT} already exists. It is the source of truth now — refusing to overwrite.\n"
             f"Pass --force only if you mean to discard every edit made in it.")

# ── Read the current code so the first version matches what ships ───────────
presets = open(os.path.join(REPO, "services/aiVision/stylePresets.ts"), encoding="utf-8").read()
prompts = open(os.path.join(REPO, "services/aiVision/promptTemplates.ts"), encoding="utf-8").read()
vision = open(os.path.join(REPO, "src/components/VisionExperience.tsx"), encoding="utf-8").read()

briefs_block = presets.split("export const STYLE_BRIEFS")[1].split("// ────")[0]
BRIEFS = dict(re.findall(r"^\s{2}(\w+):\s*`\n(.*?)\n`\.trim\(\),", briefs_block, re.S | re.M))
NAME_MAP = dict(re.findall(r'"([^"]+)":\s*"(\w+)",', presets.split("STYLE_NAME_TO_PRESET")[1].split("};")[0]))
PRESET_TO_NAME = {v: k for k, v in NAME_MAP.items()}
ROOM_LABELS = dict(re.findall(r'(\w+):\s*"([^"]+)",', presets.split("ROOM_TYPE_LABELS")[1].split("};")[0]))
ROOM_MAP = dict(re.findall(r'"([^"]+)":\s*"(\w+)",', presets.split("ROOM_NAME_TO_TYPE")[1].split("};")[0]))
PROGRAMS = dict(re.findall(r"^\s{2}(\w+):\s*`(.*?)`,",
                           prompts.split("ROOM_PROGRAM_RULES")[1].split("\n};")[0], re.S | re.M))
UI_STYLES = re.findall(r"'([^']+)'", vision.split("VISION_STYLES_FULL = [")[1].split("] as const")[0])
UI_ROOMS = re.findall(r"'([^']+)'", vision.split("ROOM_TYPES_FULL = [")[1].split("] as const")[0])

SECTIONS = ["COLOR PALETTE", "COLOUR PALETTE", "MATERIALS & FINISHES", "FURNITURE CHARACTER",
            "LIGHTING", "WALL & CEILING TREATMENT", "DECOR & STYLING", "OVERALL MOOD"]


def split_sections(text):
    out = []
    for part in re.split(r"^\d\.\s+", text, flags=re.M)[1:]:
        part = part.strip()
        for s in SECTIONS:
            if part.upper().startswith(s):
                part = part[len(s):].lstrip(": ").strip()
                break
        out.append(part)
    while len(out) < 7:
        out.append("")
    return out[:7]


# ── Styling ─────────────────────────────────────────────────────────────────
NAVY, TERRA, AMBER = "0B2240", "9E5E41", "FFF4E0"
hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
hfill = PatternFill("solid", fgColor=NAVY)
wrap = Alignment(wrap_text=True, vertical="top")
top = Alignment(vertical="top")
ctr = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def head(ws, cols, height=30):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font, cell.fill, cell.alignment = hdr, hfill, ctr
    ws.row_dimensions[1].height = height


def body(ws, row, ncols, height):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.alignment, cell.border, cell.font = wrap, thin, Font(size=9)


wb = Workbook()

# ── README ──────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "README"
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 120
rows = [
    ("AI VISION — THE RULEBOOK", ""),
    ("", ""),
    ("THIS FILE GENERATES THE PROMPT",
     r"Run  python scripts\aivision\compile-rulebook.py  after editing. It reads every sheet below and "
     r"regenerates services\aiVision\rulebook.generated.ts, which the prompt builder imports. Edit here and "
     "the prompt changes. Never hand-edit the generated file — that is how a rule quietly stops being a rule."),
    ("One file, five sheets",
     "Rules — what the tool may and may not do to a room.\n"
     "Style Briefs — the 7-section description behind every style chip.\n"
     "Palettes — nine paint colours per style; ONE accent is used per generation.\n"
     "Room Programs — what furniture each room type must contain.\n"
     "Paint 2026 — the Colour-of-the-Year modifier that overrides the accent."),
    ("Replaces",
     "Redesign-My-Room-Rulebook.xlsx and AI-Vision-Style-Briefs.xlsx. Both are superseded by this file; "
     "keep them on disk only until the aivision-structure-bench branch is merged, then delete."),
    ("", ""),
    ("HOW A PROMPT IS ASSEMBLED", ""),
    ("Fixed every time",
     "The Rules block, the framing sentences, and the 'what to transform' list. Byte-identical for every "
     "user and every generation."),
    ("Selected, not rewritten",
     "One style brief, one room programme, one palette accent. Chosen from these sheets — never reworded "
     "per request. If the text drifted you could not tell whether a bad result came from the model or from "
     "that request's prompt, and the benchmark would stop being comparable."),
    ("Computed per request",
     "Spatial constraints measured from the uploaded photo, and the variation hint on repeat generations."),
    ("Where variety comes from",
     "The palette accent (a different one per generation) and model temperature — not from rewording."),
    ("", ""),
    ("ENFORCEMENT — read before adding a rule", ""),
    ("Prompt text does not enforce",
     "Proven twice. The prompt already banned invented architecture across 22 bullets and 10 of 16 results "
     "invented some anyway; the ceiling ban was then written out in full and the very next run invented "
     "ceiling coves. A rule stated and never checked is decoration (RD22)."),
    ("Length is not thoroughness",
     "Measured 29 Aug 2026 on the same 16 rooms, staging engine: 500+ words = 0/16 preserved (outputs "
     "unrelated to the source); 111 words = 8/16; 260 words = 6/16 and worse on every dimension. The "
     "staging rule set is empty by design (RD23)."),
    ("Numbering", "Never renumber — code and grader dimensions reference the IDs. To retire a rule set "
                  "Status to Removed; never delete the row."),
]
for i, (a, b) in enumerate(rows, start=1):
    ws.cell(i, 1, a).alignment = top
    ws.cell(i, 2, b).alignment = wrap
ws["A1"].font = Font(size=14, bold=True, color=NAVY)
for r in (3, 4, 5, 8, 9, 10, 11, 14, 15, 16):
    ws.cell(r, 1).font = Font(bold=True, size=10, color=TERRA)
    ws.row_dimensions[r].height = 56
for r in (7, 13):
    ws.cell(r, 1).font = Font(bold=True, size=11, color=NAVY)
ws.row_dimensions[4].height = 86
ws.row_dimensions[3].height = 70

# ── Rules (carried over verbatim) ───────────────────────────────────────────
ws = wb.create_sheet("Rules")
if os.path.exists(OLD_RULES):
    src = load_workbook(OLD_RULES, data_only=True)["Rules"]
    for r, row in enumerate(src.iter_rows(values_only=True), start=1):
        ws.append(list(row))
        if r == 1:
            for c in range(1, len(row) + 1):
                cell = ws.cell(1, c)
                cell.font, cell.fill, cell.alignment = hdr, hfill, ctr
            ws.row_dimensions[1].height = 30
        else:
            body(ws, r, len(row), 74)
    for c, w in zip("ABCDEFGHIJ", [8, 14, 62, 12, 40, 30, 10, 78, 11, 30]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "C2"
else:
    print(f"WARNING: {OLD_RULES} not found — Rules sheet left empty.")

# ── Style Briefs ────────────────────────────────────────────────────────────
ws = wb.create_sheet("Style Briefs")
cols = ["Style (UI chip)", "Preset key", "Live in UI?",
        "1 · Colour palette", "2 · Materials & finishes", "3 · Furniture character",
        "4 · Lighting", "5 · Wall & ceiling", "6 · Decor & styling", "7 · Overall mood"]
head(ws, cols)
for key, text in BRIEFS.items():
    label = PRESET_TO_NAME.get(key, key.replace("_", " ").title())
    ws.append([label, key, "yes" if label in UI_STYLES else "NO — unreachable"] + split_sections(text))
    body(ws, ws.max_row, len(cols), 132)
    ws.cell(ws.max_row, 1).font = Font(size=10, bold=True, color=NAVY)
for c, w in zip("ABCDEFGHIJ", [20, 18, 14, 52, 52, 52, 46, 46, 52, 46]):
    ws.column_dimensions[c].width = w
ws.freeze_panes = "D2"

# ── Palettes ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Palettes")
cols = ["Style (UI chip)", "Preset key", "#", "Colour name", "Hex", "Role", "Swatch"]
head(ws, cols)
for key, colours in PALETTES.items():
    label = PRESET_TO_NAME.get(key, key)
    for i, (name, hexv, role) in enumerate(colours, start=1):
        ws.append([label if i == 1 else "", key if i == 1 else "", i, name, hexv, role, ""])
        r = ws.max_row
        body(ws, r, len(cols), 20)
        ws.cell(r, 5).font = Font(size=9, name="Consolas")
        # The swatch cell IS the colour — the point is to review this by eye.
        ws.cell(r, 7).fill = PatternFill("solid", fgColor=hexv.lstrip("#").upper())
        if role == "accent":
            ws.cell(r, 6).font = Font(size=9, bold=True, color=TERRA)
        if i == 1:
            ws.cell(r, 1).font = Font(size=10, bold=True, color=NAVY)
for c, w in zip("ABCDEFG", [20, 18, 5, 26, 11, 11, 16]):
    ws.column_dimensions[c].width = w
ws.freeze_panes = "A2"

# ── Room Programs ───────────────────────────────────────────────────────────
ws = wb.create_sheet("Room Programs")
cols = ["Room (UI chip)", "Room key", "Prompt label", "Live in UI?", "Programme rule handed to the model"]
head(ws, cols)
key_to_ui = {}
for lbl, k in ROOM_MAP.items():
    key_to_ui.setdefault(k, []).append(lbl)
for key, label in ROOM_LABELS.items():
    uis = [u for u in key_to_ui.get(key, []) if u in UI_ROOMS]
    ws.append([" / ".join(uis) if uis else "—", key, label, "yes" if uis else "NO",
               PROGRAMS.get(key, "").strip()])
    body(ws, ws.max_row, len(cols), 118)
    ws.cell(ws.max_row, 1).font = Font(size=10, bold=True, color=NAVY)
for c, w in zip("ABCDE", [22, 16, 30, 12, 140]):
    ws.column_dimensions[c].width = w
ws.freeze_panes = "E2"

# ── Paint 2026 ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("Paint 2026")
cols = ["Id", "Colour", "Brand / code", "Hex", "Swatch", "Role", "What it is", "Instruction sent to the model"]
head(ws, cols)
for p in PAINT_2026:
    ws.append([p["id"], p["name"], p["brand"], p["hex"], "", p["role"], p["note"], p["instruction"]])
    r = ws.max_row
    body(ws, r, len(cols), 54)
    ws.cell(r, 4).font = Font(size=9, name="Consolas")
    ws.cell(r, 5).fill = PatternFill("solid", fgColor=p["hex"].lstrip("#").upper())
    ws.cell(r, 2).font = Font(size=10, bold=True, color=NAVY)
ws.append([])
ws.append(["", "Hexes are close approximations read from published swatches, not brand data files. "
               "Confirm against a real fan deck before printing anything."])
ws.cell(ws.max_row, 2).font = Font(size=9, italic=True, color=TERRA)
ws.cell(ws.max_row, 2).alignment = wrap
for c, w in zip("ABCDEFGH", [16, 20, 26, 11, 12, 10, 54, 76]):
    ws.column_dimensions[c].width = w

os.makedirs(PLAN, exist_ok=True)
wb.save(OUT)
print(f"seeded {OUT}")
print(f"  rules sheet   : carried over from {os.path.basename(OLD_RULES)}")
print(f"  style briefs  : {len(BRIEFS)}")
print(f"  palettes      : {len(PALETTES)} styles x {len(next(iter(PALETTES.values())))} colours")
print(f"  room programs : {len(PROGRAMS)}")
print(f"  paint 2026    : {len(PAINT_2026)}")
