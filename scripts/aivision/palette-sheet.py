#!/usr/bin/env python
r"""
Render the mood + colour plan for every style as one reviewable page.

    python scripts/aivision/palette-sheet.py
    python scripts/aivision/palette-sheet.py "E:\somewhere\else.html"

The data already exists in AI-Vision-Rulebook.xlsx, but split: the mood is
column 7 of "Style Briefs" and the colours are a separate "Palettes" sheet, so
nobody can see a style's plan in one place. This joins them into a paint-chip
card per style — swatches, roles, the accent rotation, and the mood line the
model is actually given.

GENERATED. Edit the workbook and re-run; never edit the HTML.
"""
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

XLSX = r"E:\Business\Claude\_Plan\Website\AI-Vision-Rulebook.xlsx"
OUT = sys.argv[1] if len(sys.argv) > 1 else r"E:\Business\Claude\_Plan\Website\AI-Vision-Palettes.html"

wb = load_workbook(XLSX, data_only=True)
sb, pal, p26 = wb["Style Briefs"], wb["Palettes"], wb["Paint 2026"]

# ── Briefs: display name, preset, colour line, mood line ────────────────────
briefs = {}
order = []
for r in range(2, sb.max_row + 1):
    key = sb.cell(r, 2).value
    if not key:
        continue
    briefs[key] = {
        "label": sb.cell(r, 1).value or key,
        "colour": (sb.cell(r, 4).value or "").strip(),
        "mood": (sb.cell(r, 10).value or "").strip(),
    }
    order.append(key)

# ── Palettes: 9 colours per style, carried down from the first row ──────────
palettes, current = {}, None
for r in range(2, pal.max_row + 1):
    k = pal.cell(r, 2).value
    if k:
        current = str(k).strip()
    name = pal.cell(r, 4).value
    if not current or not name:
        continue
    palettes.setdefault(current, []).append({
        "name": str(name).strip(),
        "hex": str(pal.cell(r, 5).value or "").strip().upper(),
        "role": str(pal.cell(r, 6).value or "").strip().lower(),
    })

missing = [k for k in order if k not in palettes]
if missing:
    sys.exit(f"No palette for: {', '.join(missing)}")


def esc(s):
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def luminance(hexv):
    """Pick readable label colour on a swatch, and flag near-white chips."""
    r, g, b = (int(hexv[i:i + 2], 16) / 255 for i in (1, 3, 5))
    f = lambda c: c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4
    return 0.2126 * f(r) + 0.7152 * f(g) + 0.0722 * f(b)


cards = []
for key in order:
    b = briefs[key]
    colours = palettes[key]
    accents = [c for c in colours if c["role"] == "accent"]

    chips = []
    for c in colours:
        light = luminance(c["hex"]) > 0.7
        chips.append(
            f'<figure class="chip {c["role"]}">'
            f'<span class="disc{" pale" if light else ""}" style="background:{c["hex"]}"></span>'
            f'<figcaption><b>{esc(c["name"])}</b><code>{esc(c["hex"])}</code>'
            f'<i class="role {c["role"]}">{c["role"]}</i></figcaption></figure>'
        )

    rotation = " · ".join(
        f'<span class="rot"><b>{i}</b> {esc(a["name"])}</span>' for i, a in enumerate(accents)
    )

    cards.append(f"""
<section class="card" id="{esc(key)}">
  <header>
    <h2>{esc(b['label'])}</h2>
    <code class="key">{esc(key)}</code>
    <span class="count">{len(accents)} accents</span>
  </header>
  <div class="body">
    <div class="chips">{''.join(chips)}</div>
    <div class="words">
      <p class="mood">{esc(b['mood'])}</p>
      <h3>Palette as the model receives it</h3>
      <p class="colour">{esc(b['colour'])}</p>
      <h3>One accent per generation</h3>
      <p class="rotation">{rotation}</p>
      <p class="note">Seed 0 is the first generation; each “Generate variation” steps to the next,
      so a repeat never lands on the colour just rejected.</p>
    </div>
  </div>
</section>""")

# -- Paint 2026: the Colour-of-the-Year modifier ----------------------------
paints = []
for r in range(2, p26.max_row + 1):
    hexv = str(p26.cell(r, 4).value or "").strip().upper()
    if not hexv.startswith("#") or len(hexv) != 7:
        continue  # trailing note rows
    paints.append({
        "name": str(p26.cell(r, 2).value or "").strip(),
        "brand": str(p26.cell(r, 3).value or "").strip(),
        "hex": hexv,
        "role": str(p26.cell(r, 6).value or "").strip(),
        "note": str(p26.cell(r, 7).value or "").strip(),
        "instruction": str(p26.cell(r, 8).value or "").strip(),
    })

paint_cards = "".join(
    f'<figure class="paint">'
    f'<span class="disc{" pale" if luminance(p["hex"]) > 0.7 else ""}" style="background:{p["hex"]}"></span>'
    f'<figcaption><b>{esc(p["name"])}</b><em>{esc(p["brand"])}</em><code>{esc(p["hex"])}</code>'
    f'<i class="role accent">{esc(p["role"])}</i>'
    f'<span class="pnote">{esc(p["note"])}</span>'
    f'<span class="pinstr">{esc(p["instruction"])}</span>'
    f'</figcaption></figure>' for p in paints)

paint_section = f"""
<section class="card paint2026" id="paint-2026">
  <header>
    <h2>2026 Colour of the Year</h2>
    <code class="key">modifier</code>
    <span class="count">{len(paints)} options &middot; not a style</span>
  </header>
  <div class="body paintbody">
    <div class="paints">{paint_cards}</div>
    <div class="words">
      <p class="mood">Choosing one of these does not replace the style. It replaces the single
      ACCENT the style&rsquo;s palette would have supplied, so the colour is guaranteed to appear
      while the materials, furniture, lighting and mood all survive intact.</p>
      <h3>Why it works with reference photos too</h3>
      <p class="colour">Because it overrides the accent rather than the brief, it applies even when a
      visitor uploads their own inspiration photos and there is no preset in play at all.</p>
      <h3>Where it is applied</h3>
      <p class="colour">In <code>buildGenerationPrompt</code>, never inside the style brief. Briefs are
      cached by <code>getCacheKey</code>; folding a colour into one would cache it there and leak that
      colour into the visitor&rsquo;s next, unrelated generation.</p>
      <p class="note">Hexes are close approximations read from published swatches, not brand data
      files. Confirm against a real fan deck before printing anything.</p>
    </div>
  </div>
</section>"""

nav = '<a href="#paint-2026">Colour of the Year</a> &middot; ' + " · ".join(f'<a href="#{esc(k)}">{esc(briefs[k]["label"])}</a>' for k in order)

html = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Vision — mood &amp; colour plan, {len(order)} styles</title>
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,400;0,600;1,400&family=Montserrat:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
  :root{{--navy:#0B2240;--terra:#9E5E41;--cobalt:#0047AB;--mute:#6B6B6B;--line:rgba(0,0,0,.12)}}
  *{{box-sizing:border-box}}
  body{{margin:0;background:#fff;color:#111;font-family:Montserrat,system-ui,sans-serif}}
  header.top{{padding:34px 30px 18px;border-bottom:1px solid var(--line);position:sticky;top:0;background:#fff;z-index:5}}
  h1{{font-family:"Cormorant Garamond",Georgia,serif;font-size:34px;margin:0 0 6px;color:var(--navy)}}
  .sub{{font-size:12px;color:var(--mute);line-height:1.65;max-width:96ch}}
  .sub b{{color:var(--terra)}}
  nav{{margin-top:12px;font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;line-height:2.1}}
  nav a{{color:var(--navy);text-decoration:none;border-bottom:1px solid transparent}}
  nav a:hover{{border-color:var(--navy)}}
  main{{padding:10px 30px 80px}}

  .card{{border:1px solid var(--line);margin:26px 0;scroll-margin-top:150px;break-inside:avoid}}
  .card header{{display:flex;align-items:baseline;gap:14px;padding:15px 20px;border-bottom:1px solid var(--line);background:#FAFAFA}}
  .card h2{{font-family:"Cormorant Garamond",Georgia,serif;font-size:26px;margin:0;color:var(--navy)}}
  .key{{font-family:ui-monospace,Consolas,monospace;font-size:11px;color:var(--mute);background:#fff;border:1px solid var(--line);padding:2px 7px}}
  .count{{margin-left:auto;font-size:9px;font-weight:800;letter-spacing:.16em;text-transform:uppercase;color:var(--terra)}}

  .body{{display:grid;grid-template-columns:minmax(330px,420px) 1fr;gap:30px;padding:22px 20px 26px}}
  .chips{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px}}
  .chip{{margin:0;text-align:center}}
  .disc{{display:block;width:100%;aspect-ratio:1;border-radius:50%;box-shadow:inset 0 0 0 6px rgba(255,255,255,.55),0 1px 4px rgba(0,0,0,.16)}}
  .disc.pale{{box-shadow:inset 0 0 0 6px rgba(255,255,255,.55),0 0 0 1px rgba(0,0,0,.18),0 1px 4px rgba(0,0,0,.10)}}
  figcaption{{margin-top:7px;font-size:9.5px;line-height:1.5}}
  figcaption b{{display:block;font-weight:700;color:#222}}
  figcaption code{{display:block;font-family:ui-monospace,Consolas,monospace;font-size:9px;color:var(--mute)}}
  .role{{display:inline-block;margin-top:3px;font-style:normal;font-size:7.5px;font-weight:800;letter-spacing:.14em;
        text-transform:uppercase;color:var(--mute)}}
  .role.accent{{color:#fff;background:var(--terra);padding:2px 6px;border-radius:99px}}
  .role.field{{color:var(--cobalt)}}

  .words h3{{font-size:9px;font-weight:800;letter-spacing:.2em;text-transform:uppercase;color:var(--terra);
             margin:20px 0 6px}}
  .mood{{font-family:"Cormorant Garamond",Georgia,serif;font-size:20px;line-height:1.5;color:#2b2b2b;margin:0}}
  .colour{{font-size:11.5px;line-height:1.7;color:#3d3d3d;margin:0}}
  .rotation{{margin:0;font-size:11px;line-height:2}}
  .rot{{display:inline-block;border:1px solid var(--line);padding:4px 10px;margin:0 6px 6px 0;border-radius:99px}}
  .rot b{{color:var(--cobalt);margin-right:5px}}
  .note{{font-size:10.5px;color:var(--mute);line-height:1.6;margin:8px 0 0}}

  .paint2026 header{{background:#0B2240}}
  .paint2026 h2{{color:#fff}}
  .paint2026 .key{{background:transparent;border-color:rgba(255,255,255,.35);color:rgba(255,255,255,.75)}}
  .paint2026 .count{{color:#E8C9A0}}
  .paintbody{{grid-template-columns:minmax(340px,540px) 1fr}}
  .paints{{display:grid;grid-template-columns:repeat(3,1fr);gap:18px}}
  .paint figcaption em{{display:block;font-style:normal;font-size:8.5px;letter-spacing:.06em;color:var(--mute);margin-top:1px}}
  .pnote{{display:block;margin-top:7px;font-size:10px;line-height:1.55;color:#3d3d3d;text-align:left}}
  .pinstr{{display:block;margin-top:6px;font-size:9.5px;line-height:1.55;color:var(--mute);text-align:left;
           border-left:2px solid var(--terra);padding-left:8px}}
  @media(max-width:900px){{.body{{grid-template-columns:1fr}}.paintbody{{grid-template-columns:1fr}}}}
  @media print{{header.top{{position:static}} nav{{display:none}} .card{{page-break-inside:avoid}}}}
</style></head><body>
<header class="top">
  <h1>Mood &amp; colour plan</h1>
  <div class="sub"><b>{len(order)} styles</b> &middot; nine paint colours each, in
    <b>field</b> (large surfaces) / <b>neutral</b> (support) / <b>accent</b> (the one strong note) roles.
    Exactly one accent is used per generation, which is what stops fifteen styles producing fifteen
    versions of the same room. Generated from <code>AI-Vision-Rulebook.xlsx</code> — edit the workbook
    and re-run <code>scripts/aivision/palette-sheet.py</code>; do not edit this page.</div>
  <nav>{nav}</nav>
</header>
<main>{paint_section}{''.join(cards)}</main>
</body></html>"""

os.makedirs(os.path.dirname(OUT), exist_ok=True)
open(OUT, "w", encoding="utf-8").write(html)
print(f"wrote {OUT}")
print(f"{len(order)} styles · {sum(len(v) for v in palettes.values())} colours · "
      f"{sum(1 for v in palettes.values() for c in v if c['role'] == 'accent')} accents")
for k in order:
    n = len([c for c in palettes[k] if c["role"] == "accent"])
    if n < 3:
        print(f"  WARNING {k}: only {n} accents — generations will repeat")
