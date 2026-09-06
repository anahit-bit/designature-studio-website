#!/usr/bin/env python3
"""Export every sheet of the website plan workbook to CSV, one file per sheet.

The workbook is the master. These CSVs exist so `git diff` shows which cells
changed, which an xlsx (a zip of XML) never will. Regenerate them whenever the
workbook changes, and commit both together.

    python3 docs/plan/scripts/export_plan.py            # rewrite the CSVs
    python3 docs/plan/scripts/export_plan.py --check    # exit 1 if they are stale

Requires openpyxl (pip install openpyxl).
"""

import argparse
import csv
import datetime
import io
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Website-plan.xlsx"
EXPORT_DIR = ROOT / "export"


def slug(name: str) -> str:
    """'Strategy & Positioning' -> 'strategy-and-positioning'."""
    s = name.strip().lower().replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "sheet"


def cell_text(value) -> str:
    """One stable string per cell, so a rerun with no edits produces no diff."""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        # Dates are typed inconsistently in the workbook: some cells carry a
        # real datetime, most carry the string '2026-09-06'. Normalise to the
        # string form, and keep the time only when one was actually set.
        if (value.hour, value.minute, value.second) == (0, 0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def sheet_rows(ws):
    """Trim the trailing empty rows and columns openpyxl reports."""
    rows = [[cell_text(v) for v in row] for row in ws.iter_rows(values_only=True)]
    while rows and not any(rows[-1]):
        rows.pop()
    width = max((len(r) for r in rows), default=0)
    while width and all(len(r) < width or not r[width - 1] for r in rows):
        width -= 1
    return [r[:width] + [""] * (width - len(r)) for r in rows]


def render(rows) -> str:
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerows(rows)
    return buf.getvalue()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check",
        action="store_true",
        help="do not write; exit 1 if any CSV differs from the workbook",
    )
    args = parser.parse_args()

    if not WORKBOOK.exists():
        print("missing workbook: %s" % WORKBOOK, file=sys.stderr)
        return 2

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    EXPORT_DIR.mkdir(exist_ok=True)

    written, stale, expected = [], [], set()
    for index, ws in enumerate(wb.worksheets, 1):
        path = EXPORT_DIR / ("%02d-%s.csv" % (index, slug(ws.title)))
        expected.add(path.name)
        rows = sheet_rows(ws)
        text = render(rows)
        current = path.read_text(encoding="utf-8") if path.exists() else None
        if current == text:
            continue
        if args.check:
            stale.append(path.name)
        else:
            path.write_text(text, encoding="utf-8")
            written.append("%s (%d rows)" % (path.name, len(rows)))
    wb.close()

    orphans = sorted(
        p.name for p in EXPORT_DIR.glob("*.csv") if p.name not in expected
    )
    for name in orphans:
        if args.check:
            stale.append("%s (no such sheet)" % name)
        else:
            (EXPORT_DIR / name).unlink()
            written.append("%s removed (no such sheet)" % name)

    if args.check:
        if stale:
            print("stale export, rerun export_plan.py:")
            for name in stale:
                print("  " + name)
            return 1
        print("export matches the workbook")
        return 0

    if not written:
        print("export already matches the workbook")
    else:
        print("wrote %d file(s) to %s" % (len(written), EXPORT_DIR))
        for line in written:
            print("  " + line)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
