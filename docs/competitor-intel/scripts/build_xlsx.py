import csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/home/user/designature-studio-website/docs/competitor-intel/competitors-2026-Q3.csv"
OUT = "/home/user/designature-studio-website/docs/competitor-intel/DesignatureStudio-Competitors-2026-Q3.xlsx"

GROUP = {
 "Hybrid AI+Human e-design":"Hybrid AI + Human e-design",
 "Hybrid AI+Human e-design (outdoor)":"Hybrid AI + Human e-design",
 "Hybrid e-design (precedent)":"Defunct / precedent",
 "Shop-the-look (precedent)":"Defunct / precedent",
 "Pure AI render/redesign":"Pure AI redesign tool",
 "Pure AI render/redesign + staging":"Pure AI redesign tool",
 "Pro AI render/redesign":"Pure AI redesign tool",
 "Pro AI render":"Pure AI redesign tool",
 "Pure AI redesign + shop-the-look":"AI redesign + shoppable (the wedge)",
 "Shop-the-look / AI redesign":"AI redesign + shoppable (the wedge)",
 "Pro/CAD software w/ AI":"Pro / design software",
 "Pro/CAD software":"Pro / design software",
 "Pro/CAD software + service":"Pro / design software",
 "E-design business suite (B2B)":"Pro / design software",
 "Studio ops software (B2B)":"Pro / design software",
 "FF&E procurement software (B2B)":"Pro / design software",
 "Generative floor-plan AI":"Pro / design software",
 "Generative architecture AI":"Pro / design software",
 "Generative real-estate AI":"Pro / design software",
 "3D/VR presentation software":"Pro / design software",
 "Spatial-data API (B2B)":"Pro / design software",
 "Pro/prosumer platform + community":"Pro / design software",
 "Mobile AI design app":"Mobile app",
 "Mobile AI design app / marketplace":"Mobile app",
 "Mobile AI design app / staging":"Mobile app",
 "Mobile AR measurement app":"Mobile app",
 "Mobile DIY design app":"Mobile app",
 "AI virtual staging":"AI virtual staging",
 "AI+human virtual staging":"AI virtual staging",
 "AI virtual staging + design":"AI virtual staging",
 "Virtual staging (was shoppable 3D)":"AI virtual staging",
 "Shop-the-look (retailer AI)":"Retailer / shop-the-look commerce",
 "Shop-the-look (retailer AR)":"Retailer / shop-the-look commerce",
 "Shop-the-look (B2B visualizer)":"Retailer / shop-the-look commerce",
 "Inspiration + shoppable pins":"Retailer / shop-the-look commerce",
 "Furniture retailer free design":"Retailer free design service",
 "Furniture brand (adjacent)":"Adjacent",
 "Physical sampling (adjacent)":"Adjacent",
 "Regional AI redesign/staging":"Regional / emerging",
 "Regional 3D planner":"Regional / emerging",
 "Regional AI redesign":"Regional / emerging",
 "Regional studio + AI":"Regional / emerging",
 "Regional bespoke studio (geo peer)":"Regional / emerging",
 "Regional full-stack execution":"Regional / emerging",
 "Horizontal generative AI":"Horizontal generative AI",
}

with open(SRC) as f:
    rows = list(csv.DictReader(f))
for r in rows:
    r["Group"] = GROUP.get(r["Category"], r["Category"])

# sort: closeness desc, then group, then name
rows.sort(key=lambda r:(-int(r["Closeness (1-5)"]), r["Group"], r["Name"]))

wb = Workbook()
ws = wb.active
ws.title = "Competitors Q3 2026"

cols = ["Name","URL","HQ / Region","Group","Category","Closeness (1-5)","Status",
        "What it is","Pricing (2026)","Covers","Does NOT cover / gaps","Tech / models (where known)"]
headers = ["Competitor","Website","HQ / Region","Group","Sub-category","Closeness\n(1-5)","Status",
           "What it is","Pricing (2026)","Covers","Does NOT cover / Gaps","Tech / models\n(where known)"]

NAVY=PatternFill("solid",fgColor="0A1F44"); WHITE=Font(color="FFFFFF",bold=True,size=11,name="Calibri")
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
top=Alignment(vertical="top",wrap_text=True)

ws.append(headers)
for c in range(1,len(headers)+1):
    cell=ws.cell(1,c); cell.fill=NAVY; cell.font=WHITE
    cell.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True); cell.border=border

CLOSE_FILL={5:"C6472F",4:"E08A5B",3:"F2D493",2:"DDE7DD",1:"EDEDED",0:"3A3A3A"}
CLOSE_FONT={0:"FFFFFF"}
for r in rows:
    ws.append([r[c] for c in cols])
    i=ws.max_row
    cl=int(r["Closeness (1-5)"])
    for c in range(1,len(cols)+1):
        cell=ws.cell(i,c); cell.alignment=top; cell.border=border; cell.font=Font(size=10,name="Calibri")
    cc=ws.cell(i,6); cc.fill=PatternFill("solid",fgColor=CLOSE_FILL[cl])
    cc.font=Font(size=11,bold=True,name="Calibri",color=CLOSE_FONT.get(cl,"000000"))
    cc.alignment=Alignment(vertical="center",horizontal="center")
    # hyperlink website
    url=r["URL"]
    if url and url not in("(dead)",):
        href=url if url.startswith("http") else "https://"+url.split(";")[0].strip()
        lc=ws.cell(i,2); lc.hyperlink=href; lc.font=Font(size=10,color="0563C1",underline="single",name="Calibri")

widths=[22,26,18,26,26,10,20,40,40,44,44,40]
for idx,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(idx)].width=w
ws.freeze_panes="A2"
ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{ws.max_row}"
ws.row_dimensions[1].height=32

# Legend / README sheet
ws2=wb.create_sheet("How to read")
notes=[
 ["Designature Studio — AI Interior Design Competitor Scan",""],
 ["Quarter","Q3 2026 (data as of Aug 2026)"],
 ["Total competitors tracked",str(len(rows))],
 ["Direct competitors (closeness 4-5)",str(sum(1 for r in rows if int(r['Closeness (1-5)'])>=4))],
 ["",""],
 ["Closeness scale","5 = direct competitor (same model); 1 = distant/adjacent; 0 = defunct/inactive precedent"],
 ["Benchmark (Designature)","Free AI (AI Vision 3 concepts + Shopping List PDF) / $19 / $49 + $99 human consult; Yerevan + worldwide remote"],
 ["",""],
 ["Groups","Hybrid AI + Human e-design (closest model match)"],
 ["","AI redesign + shoppable (the wedge — head-to-head on Shopping List)"],
 ["","Pure AI redesign tool / Mobile app / AI virtual staging"],
 ["","Retailer / shop-the-look commerce (free but catalog-locked)"],
 ["","Retailer free design service / Pro design software / Regional / emerging"],
 ["",""],
 ["Tech / models column","AI models each tool runs on, filled only where publicly disclosed or founder-confirmed (most are '—' — a research field to fill over quarters). Key insight: MeltFlex uses off-the-shelf GPT + Gemini 2.5 Flash Image + Veo/Seedance = no model moat."],
 ["",""],
 ["Refresh cadence","Quarterly. Re-run the same 6-segment scan; compare vs prior quarter's CSV in git history."],
 ["Source of truth CSV","docs/competitor-intel/competitors-2026-Q3.csv"],
]
for row in notes: ws2.append(row)
ws2.column_dimensions["A"].width=32; ws2.column_dimensions["B"].width=90
ws2.cell(1,1).font=Font(bold=True,size=14)
for i in range(1,len(notes)+1):
    ws2.cell(i,1).font=Font(bold=True,size=11) if ws2.cell(i,1).value else Font()
    ws2.cell(i,2).alignment=Alignment(wrap_text=True,vertical="top")

wb.save(OUT)
print("wrote", OUT, "with", len(rows), "rows")

# also emit normalized JSON for the artifact
json.dump(rows, open("/tmp/claude-0/-home-user-designature-studio-website/ff431010-9c81-55a6-96d3-cb1a73eff886/scratchpad/competitors.json","w"), indent=0)
print("wrote competitors.json")
from collections import Counter
for k,v in Counter(r["Group"] for r in rows).most_common(): print(f"  {v:2d}  {k}")
